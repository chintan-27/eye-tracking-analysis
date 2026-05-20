"""
db/ingest/sessions.py

Reads every s{id}_sess{n}.xlsx file and populates two tables:
  - sessions           (one row per session, ~63 rows)
  - session_alertness  (one row per paradigm per session, ~315 rows)

The xlsx is a free-form questionnaire, not a structured spreadsheet.
We parse it by scanning for known key strings in column A rather than
relying on fixed row numbers, which shift across files.

Run directly to ingest:
    python -m db.ingest.sessions
"""

import openpyxl
from rich.console import Console
from db.database import get_db
from db.config import DATA_ROOT
from db.models import Session as SessionModel, SessionAlertness

console = Console()

# ---------------------------------------------------------------------------
# _read_xlsx_as_dict()
# ---------------------------------------------------------------------------
# Reads the xlsx and returns a flat dict of {stripped_key: value}.
# Column A is the label, column B is the value.
# Keys are stripped of leading/trailing whitespace so we can look them up
# reliably regardless of how many spaces the experimenter left in.

def _read_xlsx_as_dict(path: Path) -> dict:
    """
    Read only the BEFORE EXPERIMENT section — stops as soon as it hits the
    'AFTER EACH TASK' marker. This prevents later duplicate keys (like
    'Degree of alertness', which appears 6 times) from overwriting the
    before-experiment values.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = {}
    for row in ws.iter_rows(values_only=True):
        key = row[0]
        if key is not None and str(key).strip() == "AFTER EACH TASK":
            break
        val = row[1] if len(row) > 1 else None
        if key is not None:
            rows[str(key).strip()] = val
    return rows


# ---------------------------------------------------------------------------
# _parse_date()
# ---------------------------------------------------------------------------
# The date field is a datetime object in some files and a "DD/MM/YYYY" string
# in others. We normalise both to an ISO "YYYY-MM-DD" string.

def _parse_date(raw) -> str:
    if raw is None:
        return None
    if hasattr(raw, "strftime"):
        # it's already a datetime / date object
        return raw.strftime("%Y-%m-%d")
    # it's a string like "16/09/2018"
    parts = str(raw).strip().split("/")
    if len(parts) == 3:
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return str(raw)


# ---------------------------------------------------------------------------
# _parse_time()
# ---------------------------------------------------------------------------
# Time of recording is a datetime.time object. Serialise to "HH:MM".

def _parse_time(raw) -> str:
    if raw is None:
        return None
    if hasattr(raw, "strftime"):
        return raw.strftime("%H:%M")
    return str(raw).strip()


# ---------------------------------------------------------------------------
# _parse_float()
# ---------------------------------------------------------------------------
# Converts a value to float. Handles edge cases like "5.5 + 2" where a
# subject wrote their sleep as an arithmetic expression (nap + night sleep).
# Uses eval() only on our own controlled data, wrapped in a try/except.

def _parse_float(val) -> float:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        try:
            # handle simple expressions like "5.5 + 2"
            result = eval(str(val), {"__builtins__": {}})
            return float(result)
        except Exception:
            return None


# ---------------------------------------------------------------------------
# _yes_no()
# ---------------------------------------------------------------------------
# Converts "Yes"/"No" answers to 1/0. Returns None if blank.

def _yes_no(val) -> int:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s == "yes":
        return 1
    if s == "no":
        return 0
    return None


# ---------------------------------------------------------------------------
# _parse_task_blocks()
# ---------------------------------------------------------------------------
# The AFTER EACH TASK section has 5 repeated blocks, one per paradigm.
# Each block starts with a "Task name" row and spans the next 7 rows.
# Since row numbers shift across files, we scan all rows in order and
# collect blocks by detecting "Task name" entries.
#
# We read the xlsx a second time in row order (not as a dict) to preserve
# the sequential structure of the task blocks.

def _parse_task_blocks(path: Path) -> list[dict]:
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = [(str(r[0]).strip() if r[0] else "", r[1] if len(r) > 1 else None)
            for r in ws.iter_rows(values_only=True)]

    # Keys in column A that appear inside each task block (stripped)
    DISTRACTION_KEY  = "Laboratory distractions during testing"
    NOISE_KEY        = "Noise"
    INTERRUPT_KEY    = "Personnel interruptions"
    BREAKDOWN_KEY    = "Technical breakdown"
    ALERTNESS_KEY    = "Degree of alertness"
    MISSED_ANY_KEY   = "Did you miss any trial?"
    MISSED_WHICH_KEY = "If miss, which one(s)"

    blocks = []
    i = 0
    while i < len(rows):
        key, val = rows[i]
        if key == "Task name" and val is not None:
            # start of a new task block — collect the next rows into a mini-dict
            block = {"paradigm": str(val).strip()}
            j = i + 1
            # scan up to 10 rows ahead to fill in the block fields
            while j < min(i + 10, len(rows)):
                bkey, bval = rows[j]
                if bkey.startswith(DISTRACTION_KEY):
                    block["distractions"] = bval
                elif bkey == NOISE_KEY:
                    block["noise"] = str(bval).strip() if bval else "No"
                elif bkey.startswith(INTERRUPT_KEY):
                    block["interruptions"] = str(bval).strip() if bval else "No"
                elif bkey.startswith(BREAKDOWN_KEY):
                    block["breakdown"] = _yes_no(bval)
                elif bkey.startswith(ALERTNESS_KEY):
                    block["alertness"] = str(bval).strip() if bval else None
                elif bkey.startswith(MISSED_ANY_KEY):
                    block["missed_any"] = _yes_no(bval)
                elif bkey == MISSED_WHICH_KEY:
                    block["missed_which"] = str(bval).strip() if bval else None
                j += 1
            blocks.append(block)
        i += 1

    return blocks


# ---------------------------------------------------------------------------
# run()
# ---------------------------------------------------------------------------

def run():
    Session = get_db()

    # Walk every S{id}/Sess{n}/ folder and find the xlsx file
    xlsx_files = sorted(DATA_ROOT.glob("S*/Sess*/s*_sess*.xlsx"))

    with Session() as db:
        for xlsx_path in xlsx_files:
            # Derive subject_id and session_number from the folder path.
            # We trust the folder name over the xlsx contents (xlsx values
            # sometimes have errors, e.g. session number listed as 1 in Sess02).
            parts        = xlsx_path.parts   # (..., 'S01', 'Sess01', 'file.xlsx')
            subject_id   = parts[-3]         # "S01"
            sess_folder  = parts[-2]         # "Sess01"
            session_num  = int(sess_folder.replace("Sess", ""))
            session_id   = f"{subject_id}_Sess{session_num:02d}"  # "S01_Sess01"

            if db.get(SessionModel, session_id) is not None:
                console.print(f"  [dim]↷ {session_id} (already exists)[/dim]")
                continue

            rows = _read_xlsx_as_dict(xlsx_path)

            # ------------------------------------------------------------------
            # Build the Session row from the BEFORE EXPERIMENT section
            # ------------------------------------------------------------------
            session = SessionModel(
                id               = session_id,
                subject_id       = subject_id,
                session_number   = session_num,
                date             = _parse_date(rows.get("Date of recording")),
                time_of_recording = _parse_time(rows.get("Time of recording")),
                headset_number   = str(rows["Headset number"]).strip() if rows.get("Headset number") else None,
                sleep_hours      = _parse_float(rows.get("Hours sleep last night")),
                alertness_before = str(rows["Degree of alertness"]).strip() if rows.get("Degree of alertness") else None,
                caffeine         = _yes_no(rows.get("Caffeine use today")),
                tobacco          = _yes_no(rows.get("Tobacco use today")),
                medication       = _yes_no(rows.get("Medication/Drug/Alcohol use today")),
                exercise         = _yes_no(rows.get("Recent exercise")),
                hungry           = _yes_no(rows.get("Hungry at testing")),
                remarks          = str(rows["Remarks"]).strip() if rows.get("Remarks") else None,
            )
            db.add(session)

            # ------------------------------------------------------------------
            # Build one SessionAlertness row per task block,
            # plus one "Before" row for the before-experiment alertness.
            # ------------------------------------------------------------------

            # "Before" row — only alertness is populated, rest are null
            db.add(SessionAlertness(
                session_id    = session_id,
                paradigm      = "Before",
                alertness     = str(rows["Degree of alertness"]).strip() if rows.get("Degree of alertness") else None,
                noise         = None,
                interruptions = None,
                breakdown     = None,
                missed_any    = None,
                missed_which  = None,
            ))

            # After-task rows — one per paradigm
            task_blocks = _parse_task_blocks(xlsx_path)
            for block in task_blocks:
                db.add(SessionAlertness(
                    session_id    = session_id,
                    paradigm      = block.get("paradigm"),
                    alertness     = block.get("alertness"),
                    noise         = block.get("noise"),
                    interruptions = block.get("interruptions"),
                    breakdown     = block.get("breakdown"),
                    missed_any    = block.get("missed_any"),
                    missed_which  = block.get("missed_which"),
                ))

            console.print(f"  [green]✓[/green] {session_id}  [dim](5 paradigms + Before)[/dim]")

        db.commit()


if __name__ == "__main__":
    run()
